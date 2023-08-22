import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

wb = load_workbook('jobs.xlsx')
ws = wb['ZipRecruiter']


# length = ws['a1'].value
# print(length)

longest_row = 0

# for i in range(1, ws.max_row+1):
#     print("\n")
#     print(f"Row: {i} data:")

#     for j in range(1, ws.max_column+1):
#         cell_obj = ws.cell(row=i, column=j)
#         print(cell_obj.value, end=" ")


for i in range(1, ws.max_row+1):
    print("ROW: ", i)

    for j in range(1, ws.max_column+1):
        data = ws.cell(row=i, column=j)
        print(f"{data.value} ")